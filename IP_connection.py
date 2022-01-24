
import openpyxl

file =('IP_list.xlsx')
wb_obj = openpyxl.load_workbook(file)
sheet_obj = wb_obj.active
rows=sheet_obj.max_row
IP_list=[]
for i in range(rows):
    PC_name = sheet_obj.cell(row=i+2, column=1)
    IP = sheet_obj.cell(row=i + 2, column=2)
    # print(PC_name.value)
    IP_list.append(IP.value)
print(IP_list)